import json, os, glob, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

WORK = "_refmap-work"
OUT = r"C:\Users\michael\Downloads\refid-pages-tracker.xlsx"

t2c = json.load(open(f"{WORK}/template-to-category.json", encoding="utf-8"))
unmapped = json.load(open(f"{WORK}/unmapped-categories.json", encoding="utf-8"))

def family(name, template=""):
    s = (name + " " + template).lower()
    rules = [
        ("pin kit", "Case Pin Kits"), ("king kutter", "King Kutter"),
        ("cylinders", "JD Cylinder Photos"), ("zf ", "ZF Axle"),
        ("kobelco", "Kobelco"), ("cat ", "Cat"), ("komatsu", "Komatsu"),
        ("volvo", "Volvo"), ("dresser", "Dresser"), ("hitachi", "Hitachi"),
        ("ford", "Ford/NH"), ("new holland", "Ford/NH"),
        ("deere", "John Deere"), ("jd ", "John Deere"), ("jd-", "John Deere"),
        ("case", "Case"),
    ]
    for k, v in rules:
        if k in s:
            return v
    return "Other"

HEAD_FILL = PatternFill("solid", start_color="1A1A1A")
HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)

def style_sheet(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if c.font is None or not c.font.bold:
                c.font = BODY_FONT

wb = Workbook()

# ---------- Sheet 1: migrated 109 ----------
ws1 = wb.active
ws1.title = "Migrated Maps (109)"
ws1.append(["Category ID", "Category Name", "Live URL", "Old Template", "Hotspots",
            "Product Links", "Row Jumps", "Status", "Notes"])
rows1 = []
for template, cats in t2c.items():
    for cat in cats:
        cid = cat["category_id"]
        jpath = f"assets/ref-maps/{cid}.json"
        n_h = n_p = n_a = 0
        if os.path.exists(jpath):
            d = json.load(open(jpath, encoding="utf-8"))
            n_h = len(d["hotspots"])
            n_p = sum(1 for h in d["hotspots"] if h["kind"] == "product")
            n_a = sum(1 for h in d["hotspots"] if h["kind"] == "anchor")
        rows1.append([cid, cat["name"], "https://www.brokentractor.com" + (cat["url"] or ""),
                      template, n_h, n_p, n_a, "", ""])
rows1.sort(key=lambda r: r[1].lower())
for r in rows1:
    ws1.append(r)
style_sheet(ws1, [11, 45, 60, 38, 9, 12, 10, 14, 30])

# ---------- Sheet 2: to map 129 ----------
ws2 = wb.create_sheet("To Map (129)")
ws2.append(["Category ID", "Category Name", "Family", "Live URL", "Has Image",
            "Products", "REFID Rows", "Detected Style", "Marks Found", "Status", "Notes"])
rows2 = []
for c in unmapped:
    cid = c["category_id"]
    det_style, det_n = "", ""
    dpath = f"{WORK}/cv/detect/{cid}.json"
    if os.path.exists(dpath):
        d = json.load(open(dpath, encoding="utf-8"))
        det_style = d.get("style") or "none detected"
        det_n = len(d.get("circles", []))
    rows2.append([cid, c["name"], family(c["name"]),
                  "https://www.brokentractor.com" + (c["url"] or ""),
                  "Yes" if c["images"] else "NO IMAGE",
                  c["product_count"], len(c["ref_rows"]), det_style, det_n, "", ""])
rows2.sort(key=lambda r: (r[2], r[1].lower()))
for r in rows2:
    ws2.append(r)
style_sheet(ws2, [11, 45, 17, 60, 11, 9, 11, 15, 12, 14, 30])

# highlight NO IMAGE rows
warn = PatternFill("solid", start_color="FFF3C4")
for row in ws2.iter_rows(min_row=2):
    if row[4].value == "NO IMAGE":
        for c in row:
            c.fill = warn

# ---------- Sheet 3: summary ----------
ws3 = wb.create_sheet("Summary")
n1 = len(rows1) + 1
n2 = len(rows2) + 1
ws3["A1"] = "REF-ID Clickable Diagram Project — Tracker"
ws3["A1"].font = Font(name="Arial", size=13, bold=True)
data = [
    ("", ""),
    ("Migrated maps (from old custom templates)", f"=COUNTA('Migrated Maps (109)'!A2:A{n1})"),
    ("  ...verified (Status not blank)", f"=COUNTA('Migrated Maps (109)'!H2:H{n1})"),
    ("Pages still to map (live refid-table)", f"=COUNTA('To Map (129)'!A2:A{n2})"),
    ("  ...with no diagram image", f"=COUNTIF('To Map (129)'!E2:E{n2},\"NO IMAGE\")"),
    ("  ...verified (Status not blank)", f"=COUNTA('To Map (129)'!J2:J{n2})"),
    ("", ""),
    ("Suggested Status values:", "Map Ready / Verified Local / LIVE / Skip / Problem"),
]
for i, (label, val) in enumerate(data, 2):
    ws3[f"A{i}"] = label
    ws3[f"B{i}"] = val
    ws3[f"A{i}"].font = BODY_FONT
    ws3[f"B{i}"].font = BODY_FONT
ws3.column_dimensions["A"].width = 45
ws3.column_dimensions["B"].width = 45

wb.move_sheet("Summary", offset=-2)
wb.save(OUT)
print(f"Saved {OUT}")
print(f"Sheet1 rows: {len(rows1)}, Sheet2 rows: {len(rows2)}")
